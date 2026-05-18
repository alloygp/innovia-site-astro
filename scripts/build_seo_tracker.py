"""
Build the AGP SEO Tracker workbook for a client.

Usage:
  python scripts/build_seo_tracker.py "Client Name" "clientdomain.com" "Owner Name"

Example:
  python scripts/build_seo_tracker.py "Tidewater CAM" "tidewatercam.com" "Skyler"

Output: _build/{slug}_seo_tracker.xlsx
"""

import sys
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- Args ----------
if len(sys.argv) < 3:
    print("Usage: python scripts/build_seo_tracker.py \"Client Name\" \"domain.com\" [\"Owner\"]")
    sys.exit(1)

CLIENT_NAME   = sys.argv[1]
CLIENT_DOMAIN = re.sub(r"^https?://", "", sys.argv[2].strip()).rstrip("/")
OWNER         = sys.argv[3] if len(sys.argv) > 3 else "Skyler"
BASE_URL      = f"https://{CLIENT_DOMAIN}"
SLUG          = re.sub(r"[^a-z0-9]+", "_", CLIENT_NAME.lower()).strip("_")

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "_build")
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, f"{SLUG}_seo_tracker.xlsx")

# ---------- Brand palette ----------
NAVY     = "0F1E3D"
NAVY_DK  = "0A1530"
ACCENT   = "C9A961"
LIGHT_BG = "F5F7FA"
SUBHEAD  = "E5E9F2"
ROW_ALT  = "FAFBFD"
GREEN    = "D1FAE5";  GREEN_DK  = "065F46"
AMBER    = "FEF3C7";  AMBER_DK  = "92400E"
RED      = "FEE2E2";  RED_DK    = "991B1B"
GRAY     = "E5E7EB";  GRAY_DK   = "374151"
WHITE    = "FFFFFF"
BORDER_GRAY = "D1D5DB"
FONT_NAME   = "Calibri"

thin = Side(style="thin", color=BORDER_GRAY)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_style(cell):
    cell.font      = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    cell.fill      = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = box

def body_style(cell, wrap=True):
    cell.font      = Font(name=FONT_NAME, size=10, color="111827")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    cell.border    = box

def title_style(cell, size=18):
    cell.font      = Font(name=FONT_NAME, size=size, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def section_style(cell):
    cell.font      = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
    cell.fill      = PatternFill("solid", start_color=SUBHEAD)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = box

def write_headers(ws, headers, row=1):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        header_style(c)
    ws.row_dimensions[row].height = 36

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()
wb.remove(wb.active)

# =========================================================================
# 1. README / COVER
# =========================================================================
ws = wb.create_sheet("README")
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = NAVY

ws.merge_cells("B2:H2")
ws["B2"] = f"{CLIENT_NAME} — SEO Tracker"
title_style(ws["B2"], size=24)
ws.row_dimensions[2].height = 40

ws.merge_cells("B3:H3")
ws["B3"] = "Operational source of truth for on-page, technical, link, and performance work"
ws["B3"].font = Font(name=FONT_NAME, size=11, italic=True, color=GRAY_DK)
ws.row_dimensions[3].height = 20

ws.merge_cells("B4:H4")
ws["B4"] = f"Site: {CLIENT_DOMAIN}  •  Owner: {OWNER}  •  AGP Template v1.0"
ws["B4"].font = Font(name=FONT_NAME, size=10, color=GRAY_DK)

ws["B6"] = "How to use this workbook"
section_style(ws["B6"])
ws.merge_cells("B6:H6")
ws.row_dimensions[6].height = 22

how = [
    ("1. Page Inventory",  "Master list of every URL. Capture page type, target keyword, current and proposed title/H1/meta, schema, and status. Title/description/H1 length cells auto-flag overruns."),
    ("2. Keyword Targets", "Keyword-first view independent of pages. Tracks volume, difficulty, intent, current rank, target rank. Opportunity score auto-calculated to sort priorities."),
    ("3. Technical Health","Per-URL technical checks: indexability, schema, mobile, CWV snapshot, internal link count, last crawl. Pair with Search Console data."),
    ("4. Internal Linking","Source → Target edge list. Use COUNTIF to spot orphans (0 inbound) and over-pointed pages."),
    ("5. Performance",     "Monthly traffic, top queries, conversions, and trend per page. One row per page per month."),
    ("6. Backlinks",       "Optional — referring domain log. Track only acquired/notable links, not the full Ahrefs export."),
    ("7. Roadmap",         "Action layer. Every task ties to a page (or 'site-wide'), has a priority, owner, due date, and status."),
    ("8. Lookups",         "Hidden-style reference lists that drive dropdowns. Edit here to change validation values everywhere."),
]
r = 7
for label, desc in how:
    ws[f"B{r}"] = label
    ws[f"B{r}"].font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
    ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(f"C{r}:H{r}")
    ws[f"C{r}"] = desc
    ws[f"C{r}"].font = Font(name=FONT_NAME, size=10, color="111827")
    ws[f"C{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
ws.merge_cells(f"B{r}:H{r}")
ws[f"B{r}"] = "Conventions"
section_style(ws[f"B{r}"])
ws.row_dimensions[r].height = 22
r += 1

conv = [
    ("URLs",                   "Always full https URLs. One row per canonical URL. Use trailing slashes consistently with the live site."),
    ("Title length target",    "50–60 characters. <30 underuses real estate; >60 truncates in SERP."),
    ("Meta description target","120–155 characters. <80 underuses; >155 truncates."),
    ("H1 target",              "≤70 characters. One H1 per page."),
    ("Status values",          "Not started → In progress → In review → Live → Needs refresh. Set in Lookups tab."),
    ("Priority values",        "P0 (this week), P1 (this month), P2 (this quarter), P3 (backlog)."),
    ("Intent values",          "Informational, Commercial, Transactional, Navigational, Local."),
    ("Owners",                 "Use first names or initials. Keep consistent so filters work."),
    ("Dates",                  "ISO format YYYY-MM-DD."),
]
for label, desc in conv:
    ws[f"B{r}"] = label
    ws[f"B{r}"].font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
    ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(f"C{r}:H{r}")
    ws[f"C{r}"] = desc
    ws[f"C{r}"].font = Font(name=FONT_NAME, size=10, color="111827")
    ws[f"C{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 1

r += 1
ws.merge_cells(f"B{r}:H{r}")
ws[f"B{r}"] = "Color key"
section_style(ws[f"B{r}"])
ws.row_dimensions[r].height = 22
r += 1
for fill, fg, label in [
    (GREEN,   GREEN_DK, "Within target / Live / Done"),
    (AMBER,   AMBER_DK, "Caution / In progress / Needs review"),
    (RED,     RED_DK,   "Out of range / Blocked / Missing"),
    (GRAY,    GRAY_DK,  "Not started / Inactive"),
    (SUBHEAD, NAVY,     "Section / informational"),
]:
    ws[f"B{r}"] = "  "
    ws[f"B{r}"].fill = PatternFill("solid", start_color=fill)
    ws[f"B{r}"].border = box
    ws.merge_cells(f"C{r}:H{r}")
    ws[f"C{r}"] = label
    ws[f"C{r}"].font = Font(name=FONT_NAME, size=10, color=fg, bold=True)
    ws[f"C{r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 90

# =========================================================================
# 2. PAGE INVENTORY
# =========================================================================
ws = wb.create_sheet("Page Inventory")
ws.sheet_properties.tabColor = ACCENT

inv_headers = [
    "URL","Page Type","Status","Owner","Last Updated",
    "Primary Keyword","Search Intent","Secondary Keywords",
    "Current Title","Proposed Title","Title Len",
    "Current Meta Description","Proposed Meta Description","Meta Len",
    "Current H1","Proposed H1","H1 Len",
    "Schema Type","Canonical","OG Image","Notes",
]
write_headers(ws, inv_headers)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(inv_headers))}1"

example_rows = [
    [f"{BASE_URL}/",        "Homepage", "Not started", OWNER, "", "", "Commercial", "", "", "", "", "", "", "", "", "", "", "Organization",     f"{BASE_URL}/",        "", "Front door — update with positioning, proof, and CTA."],
    [f"{BASE_URL}/services/","Service Hub","Not started",OWNER,"","","Commercial","","","","","","","","","","","Service","","","Hub linking to all individual services."],
    [f"{BASE_URL}/about/",  "About",    "Not started", OWNER, "", "", "Navigational","","","","","","","","","","","AboutPage","","",""],
    [f"{BASE_URL}/contact/","Contact",  "Not started", OWNER, "", "", "Navigational","","","","","","","","","","","ContactPage","","",""],
]

DATA_ROWS  = 200
total_rows = len(example_rows) + DATA_ROWS

for i in range(total_rows):
    er = i + 2
    row_data = example_rows[i] if i < len(example_rows) else [""] * len(inv_headers)
    for j, val in enumerate(row_data, 1):
        c = ws.cell(row=er, column=j, value=val)
        body_style(c, wrap=True)
    ws.cell(row=er, column=11).value = f'=IF(J{er}="","",LEN(J{er}))'
    ws.cell(row=er, column=14).value = f'=IF(M{er}="","",LEN(M{er}))'
    ws.cell(row=er, column=17).value = f'=IF(P{er}="","",LEN(P{er}))'
    for col in (11, 14, 17):
        cc = ws.cell(row=er, column=col)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.font = Font(name=FONT_NAME, size=10, bold=True, color=GRAY_DK)
        cc.border = box
    ws.row_dimensions[er].height = 32

set_widths(ws, [44,14,14,12,13, 26,14,28, 34,34,8, 40,40,8, 30,30,8, 16,30,22,40])
last_inv = total_rows + 1

for rng, ok_lo, ok_hi, bad_hi in [
    (f"K2:K{last_inv}", 50, 60, 60),
]:
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(K2<>"",K2>60)'],   fill=PatternFill("solid",start_color=RED),  font=Font(color=RED_DK,  bold=True)))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(K2<>"",K2>=50,K2<=60)'], fill=PatternFill("solid",start_color=GREEN),font=Font(color=GREEN_DK,bold=True)))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(K2<>"",K2<50)'],   fill=PatternFill("solid",start_color=AMBER),font=Font(color=AMBER_DK,bold=True)))

ws.conditional_formatting.add(f"N2:N{last_inv}", FormulaRule(formula=[f'AND(N2<>"",N2>155)'],        fill=PatternFill("solid",start_color=RED),  font=Font(color=RED_DK,  bold=True)))
ws.conditional_formatting.add(f"N2:N{last_inv}", FormulaRule(formula=[f'AND(N2<>"",N2>=120,N2<=155)'],fill=PatternFill("solid",start_color=GREEN),font=Font(color=GREEN_DK,bold=True)))
ws.conditional_formatting.add(f"N2:N{last_inv}", FormulaRule(formula=[f'AND(N2<>"",N2<120)'],         fill=PatternFill("solid",start_color=AMBER),font=Font(color=AMBER_DK,bold=True)))
ws.conditional_formatting.add(f"Q2:Q{last_inv}", FormulaRule(formula=[f'AND(Q2<>"",Q2>70)'],  fill=PatternFill("solid",start_color=RED),  font=Font(color=RED_DK,  bold=True)))
ws.conditional_formatting.add(f"Q2:Q{last_inv}", FormulaRule(formula=[f'AND(Q2<>"",Q2<=70,Q2>0)'], fill=PatternFill("solid",start_color=GREEN),font=Font(color=GREEN_DK,bold=True)))

for val, fill, fg in [("Live",GREEN,GREEN_DK),("Done",GREEN,GREEN_DK),("In progress",AMBER,AMBER_DK),("In review",AMBER,AMBER_DK),("Blocked",RED,RED_DK),("Needs refresh",RED,RED_DK),("Not started",GRAY,GRAY_DK)]:
    ws.conditional_formatting.add(f"C2:C{last_inv}", CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",start_color=fill),font=Font(color=fg,bold=True)))

# =========================================================================
# 3. KEYWORD TARGETS
# =========================================================================
ws = wb.create_sheet("Keyword Targets")
ws.sheet_properties.tabColor = ACCENT
kw_headers = ["Keyword","Search Intent","Volume","Difficulty (KD)","Current Rank","Target Rank","Target Page URL","Status","Priority","Opportunity Score","SERP Features","Last Checked","Notes"]
write_headers(ws, kw_headers)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(kw_headers))}1"

KW_ROWS = 250
for i in range(KW_ROWS):
    er = i + 2
    for j in range(1, len(kw_headers)+1):
        c = ws.cell(row=er, column=j, value="")
        body_style(c, wrap=True)
    ws.cell(row=er, column=10).value = f'=IF(OR(C{er}="",D{er}=""),"",ROUND(C{er}*(101-D{er})/100,0))'
    cc = ws.cell(row=er, column=10)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    cc.font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
    cc.border = box
    for col in (3,4,5,6):
        ws.cell(row=er, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[er].height = 24

set_widths(ws, [32,14,10,12,11,11,38,14,10,14,22,13,36])
last_kw = KW_ROWS + 1
ws.conditional_formatting.add(f"J2:J{last_kw}", ColorScaleRule(start_type="min",start_color="F8FAFC",mid_type="percentile",mid_value=50,mid_color="FDE68A",end_type="max",end_color="65A30D"))
ws.conditional_formatting.add(f"D2:D{last_kw}", ColorScaleRule(start_type="num",start_value=0,start_color="65A30D",mid_type="num",mid_value=40,mid_color="FDE68A",end_type="num",end_value=80,end_color="EF4444"))
for val, fill, fg in [("Live",GREEN,GREEN_DK),("Done",GREEN,GREEN_DK),("In progress",AMBER,AMBER_DK),("Blocked",RED,RED_DK),("Not started",GRAY,GRAY_DK)]:
    ws.conditional_formatting.add(f"H2:H{last_kw}", CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",start_color=fill),font=Font(color=fg,bold=True)))

# =========================================================================
# 4. TECHNICAL HEALTH
# =========================================================================
ws = wb.create_sheet("Technical Health")
ws.sheet_properties.tabColor = ACCENT
th_headers = ["URL","Indexable?","Robots Tag","Canonical OK?","Schema Type","Schema Valid?","Mobile Friendly?","LCP (s)","INP (ms)","CLS","Internal Links (in)","Internal Links (out)","Last Crawled","Last Indexed","Notes"]
write_headers(ws, th_headers)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(th_headers))}1"

TH_ROWS = 200
for i in range(TH_ROWS):
    er = i + 2
    for j in range(1, len(th_headers)+1):
        c = ws.cell(row=er, column=j, value="")
        body_style(c, wrap=True)
        if j in (2,4,6,7,8,9,10,11,12,13,14):
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[er].height = 22

set_widths(ws, [44,12,16,13,18,13,14,10,10,10,14,14,14,14,30])
last_th = TH_ROWS + 1
for col_letter in ["B","D","F","G"]:
    rng = f"{col_letter}2:{col_letter}{last_th}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal",formula=['"Yes"'],fill=PatternFill("solid",start_color=GREEN),font=Font(color=GREEN_DK,bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal",formula=['"No"'], fill=PatternFill("solid",start_color=RED),  font=Font(color=RED_DK,  bold=True)))

ws.conditional_formatting.add(f"H2:H{last_th}", FormulaRule(formula=['AND(H2<>"",H2<=2.5)'],      fill=PatternFill("solid",start_color=GREEN),font=Font(color=GREEN_DK,bold=True)))
ws.conditional_formatting.add(f"H2:H{last_th}", FormulaRule(formula=['AND(H2<>"",H2>2.5,H2<=4)'], fill=PatternFill("solid",start_color=AMBER),font=Font(color=AMBER_DK,bold=True)))
ws.conditional_formatting.add(f"H2:H{last_th}", FormulaRule(formula=['AND(H2<>"",H2>4)'],          fill=PatternFill("solid",start_color=RED),  font=Font(color=RED_DK,  bold=True)))
ws.conditional_formatting.add(f"I2:I{last_th}", FormulaRule(formula=['AND(I2<>"",I2<=200)'],       fill=PatternFill("solid",start_color=GREEN),font=Font(color=GREEN_DK,bold=True)))
ws.conditional_formatting.add(f"I2:I{last_th}", FormulaRule(formula=['AND(I2<>"",I2>200,I2<=500)'],fill=PatternFill("solid",start_color=AMBER),font=Font(color=AMBER_DK,bold=True)))
ws.conditional_formatting.add(f"I2:I{last_th}", FormulaRule(formula=['AND(I2<>"",I2>500)'],         fill=PatternFill("solid",start_color=RED),  font=Font(color=RED_DK,  bold=True)))
ws.conditional_formatting.add(f"J2:J{last_th}", FormulaRule(formula=['AND(J2<>"",J2<=0.1)'],       fill=PatternFill("solid",start_color=GREEN),font=Font(color=GREEN_DK,bold=True)))
ws.conditional_formatting.add(f"J2:J{last_th}", FormulaRule(formula=['AND(J2<>"",J2>0.1,J2<=0.25)'],fill=PatternFill("solid",start_color=AMBER),font=Font(color=AMBER_DK,bold=True)))
ws.conditional_formatting.add(f"J2:J{last_th}", FormulaRule(formula=['AND(J2<>"",J2>0.25)'],        fill=PatternFill("solid",start_color=RED),  font=Font(color=RED_DK,  bold=True)))
ws.conditional_formatting.add(f"K2:K{last_th}", CellIsRule(operator="equal",formula=["0"],fill=PatternFill("solid",start_color=RED),font=Font(color=RED_DK,bold=True)))

# =========================================================================
# 5. INTERNAL LINKING
# =========================================================================
ws = wb.create_sheet("Internal Linking")
ws.sheet_properties.tabColor = ACCENT
il_headers = ["Source URL","Target URL","Anchor Text","Context / Section","Link Type","Reviewed By","Date Added","Notes"]
write_headers(ws, il_headers)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(il_headers))}1"
IL_ROWS = 400
for i in range(IL_ROWS):
    er = i + 2
    for j in range(1, len(il_headers)+1):
        c = ws.cell(row=er, column=j, value="")
        body_style(c, wrap=True)
    ws.row_dimensions[er].height = 20
set_widths(ws, [42,42,26,26,14,14,13,30])

# =========================================================================
# 6. PERFORMANCE
# =========================================================================
ws = wb.create_sheet("Performance")
ws.sheet_properties.tabColor = ACCENT
perf_headers = ["Month (YYYY-MM)","URL","Clicks","Impressions","Avg Position","CTR (%)","Sessions (GA4)","Conversions","Conv. Rate (%)","Top Query","Top Query Clicks","Trend vs. Prev","Notes"]
write_headers(ws, perf_headers)
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(perf_headers))}1"
PERF_ROWS = 400
for i in range(PERF_ROWS):
    er = i + 2
    for j in range(1, len(perf_headers)+1):
        c = ws.cell(row=er, column=j, value="")
        body_style(c, wrap=True)
        if j in (3,4,5,6,7,8,9,11): c.alignment = Alignment(horizontal="center",vertical="center")
    ws.cell(row=er, column=6).value = f'=IF(OR(C{er}="",D{er}="",D{er}=0),"",ROUND(C{er}/D{er}*100,2))'
    ws.cell(row=er, column=9).value = f'=IF(OR(G{er}="",H{er}="",G{er}=0),"",ROUND(H{er}/G{er}*100,2))'
    for col in (6,9):
        cc = ws.cell(row=er, column=col)
        cc.font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
        cc.border = box
        cc.alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[er].height = 22
set_widths(ws, [16,42,10,13,12,10,14,13,13,26,14,14,28])
last_perf = PERF_ROWS + 1
ws.conditional_formatting.add(f"E2:E{last_perf}", FormulaRule(formula=['AND(E2<>"",E2<=3)'],      fill=PatternFill("solid",start_color=GREEN),font=Font(color=GREEN_DK,bold=True)))
ws.conditional_formatting.add(f"E2:E{last_perf}", FormulaRule(formula=['AND(E2<>"",E2>3,E2<=10)'],fill=PatternFill("solid",start_color=AMBER),font=Font(color=AMBER_DK,bold=True)))
ws.conditional_formatting.add(f"E2:E{last_perf}", FormulaRule(formula=['AND(E2<>"",E2>10)'],       fill=PatternFill("solid",start_color=RED),  font=Font(color=RED_DK,  bold=True)))

# =========================================================================
# 7. BACKLINKS
# =========================================================================
ws = wb.create_sheet("Backlinks")
ws.sheet_properties.tabColor = ACCENT
bl_headers = ["Date Acquired","Referring Domain","Referring URL","Target URL","Anchor Text","DR","Link Type","Follow / Nofollow","Source Type","Notes"]
write_headers(ws, bl_headers)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(bl_headers))}1"
BL_ROWS = 200
for i in range(BL_ROWS):
    er = i + 2
    for j in range(1, len(bl_headers)+1):
        c = ws.cell(row=er, column=j, value="")
        body_style(c, wrap=True)
        if j in (6,): c.alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[er].height = 22
set_widths(ws, [13,26,42,42,26,8,14,14,16,30])
last_bl = BL_ROWS + 1
ws.conditional_formatting.add(f"F2:F{last_bl}", ColorScaleRule(start_type="num",start_value=0,start_color="F1F5F9",mid_type="num",mid_value=40,mid_color="FDE68A",end_type="num",end_value=90,end_color="65A30D"))

# =========================================================================
# 8. ROADMAP
# =========================================================================
ws = wb.create_sheet("Roadmap")
ws.sheet_properties.tabColor = ACCENT
rm_headers = ["Task","Page / Scope","Type","Priority","Status","Owner","Due Date","Started","Completed","Notes"]
write_headers(ws, rm_headers)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(rm_headers))}1"

starter_tasks = [
    ["Audit current titles & metas across all live URLs", "site-wide","On-page","P0","Not started",OWNER,"","","","Use Page Inventory tab as the system of record."],
    ["Implement Organization schema on homepage", f"{BASE_URL}/","Technical","P0","Not started",OWNER,"","","",""],
    ["Set up GSC + GA4 monthly snapshot into Performance tab","site-wide","Reporting","P1","Not started",OWNER,"","","",""],
    ["Crawl with Screaming Frog and populate Technical Health tab","site-wide","Technical","P1","Not started",OWNER,"","","",""],
    ["Map internal links on top 10 pages","site-wide","Internal links","P1","Not started",OWNER,"","","",""],
]
RM_ROWS = 150
for i in range(RM_ROWS):
    er = i + 2
    row_data = starter_tasks[i] if i < len(starter_tasks) else [""] * len(rm_headers)
    for j, val in enumerate(row_data, 1):
        c = ws.cell(row=er, column=j, value=val)
        body_style(c, wrap=True)
    ws.row_dimensions[er].height = 28
set_widths(ws, [44,36,14,10,14,12,12,12,12,36])
last_rm = RM_ROWS + 1
for val, fill, fg in [("P0",RED,RED_DK),("P1",AMBER,AMBER_DK),("P2",GRAY,GRAY_DK),("P3",LIGHT_BG,GRAY_DK)]:
    ws.conditional_formatting.add(f"D2:D{last_rm}", CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",start_color=fill),font=Font(color=fg,bold=True)))
for val, fill, fg in [("Live",GREEN,GREEN_DK),("Done",GREEN,GREEN_DK),("In progress",AMBER,AMBER_DK),("In review",AMBER,AMBER_DK),("Blocked",RED,RED_DK),("Not started",GRAY,GRAY_DK)]:
    ws.conditional_formatting.add(f"E2:E{last_rm}", CellIsRule(operator="equal",formula=[f'"{val}"'],fill=PatternFill("solid",start_color=fill),font=Font(color=fg,bold=True)))
ws.conditional_formatting.add(f"G2:G{last_rm}", FormulaRule(formula=[f'AND(G2<>"",G2<TODAY(),NOT(OR(E2="Done",E2="Live")))'],fill=PatternFill("solid",start_color=RED),font=Font(color=RED_DK,bold=True)))

# =========================================================================
# 9. LOOKUPS
# =========================================================================
ws = wb.create_sheet("Lookups")
ws.sheet_properties.tabColor = NAVY_DK
lookups = {
    "A":("Page Type",["Homepage","Service Hub","Service","Sub-Service","Landing Page","Case Study","Hub","Blog Hub","Blog","Resource","About","Contact","Location","Legal / Utility","Other"]),
    "B":("Status",["Not started","In progress","In review","Live","Needs refresh","Blocked","Done"]),
    "C":("Priority",["P0","P1","P2","P3"]),
    "D":("Search Intent",["Informational","Commercial","Transactional","Navigational","Local"]),
    "E":("Yes / No",["Yes","No","N/A"]),
    "F":("Schema Types",["Organization","LocalBusiness","Service","Article","BlogPosting","FAQPage","BreadcrumbList","WebPage","AboutPage","ContactPage","Person","Product","Review","VideoObject","CollectionPage","ItemList","Blog"]),
    "G":("Task Type",["On-page","Technical","Content","Internal links","Schema","Reporting","Outreach","Local SEO","CRO","Other"]),
    "H":("Trend vs. Prev",["▲ Up","Flat","▼ Down","New"]),
    "I":("Link Type",["Nav","Body","Footer","Sidebar","CTA","Card","Breadcrumb"]),
    "J":("Backlink Source Type",["Earned","Outreach","Guest post","Directory","Press / PR","Partnership","Resource page","Brand mention link","Other"]),
    "K":("Follow / Nofollow",["Follow","Nofollow","UGC","Sponsored"]),
}
for col, (label, items) in lookups.items():
    cell = ws[f"{col}1"]
    cell.value = label
    header_style(cell)
    for i, item in enumerate(items, 2):
        c = ws[f"{col}{i}"]
        c.value = item
        body_style(c, wrap=False)
set_widths(ws, [22]*len(lookups))
ws.row_dimensions[1].height = 30

def lr(col, count): return f"Lookups!${col}$2:${col}${count+1}"

sheet_last = {"Page Inventory":last_inv,"Keyword Targets":last_kw,"Technical Health":last_th,"Internal Linking":IL_ROWS+1,"Performance":last_perf,"Backlinks":last_bl,"Roadmap":last_rm}

all_dv = [
    ("Page Inventory","B",lr("A",len(lookups["A"][1])),"Page Type"),
    ("Page Inventory","C",lr("B",len(lookups["B"][1])),"Status"),
    ("Page Inventory","G",lr("D",len(lookups["D"][1])),"Search Intent"),
    ("Page Inventory","R",lr("F",len(lookups["F"][1])),"Schema Type"),
    ("Keyword Targets","B",lr("D",len(lookups["D"][1])),"Search Intent"),
    ("Keyword Targets","H",lr("B",len(lookups["B"][1])),"Status"),
    ("Keyword Targets","I",lr("C",len(lookups["C"][1])),"Priority"),
    ("Technical Health","B",lr("E",len(lookups["E"][1])),"Indexable"),
    ("Technical Health","D",lr("E",len(lookups["E"][1])),"Canonical OK"),
    ("Technical Health","E",lr("F",len(lookups["F"][1])),"Schema Type"),
    ("Technical Health","F",lr("E",len(lookups["E"][1])),"Schema Valid"),
    ("Technical Health","G",lr("E",len(lookups["E"][1])),"Mobile Friendly"),
    ("Internal Linking","E",lr("I",len(lookups["I"][1])),"Link Type"),
    ("Performance","L",lr("H",len(lookups["H"][1])),"Trend"),
    ("Backlinks","G",lr("I",len(lookups["I"][1])),"Link Type"),
    ("Backlinks","H",lr("K",len(lookups["K"][1])),"Follow"),
    ("Backlinks","I",lr("J",len(lookups["J"][1])),"Source Type"),
    ("Roadmap","C",lr("G",len(lookups["G"][1])),"Task Type"),
    ("Roadmap","D",lr("C",len(lookups["C"][1])),"Priority"),
    ("Roadmap","E",lr("B",len(lookups["B"][1])),"Status"),
]
for sname, col, src, label in all_dv:
    target = wb[sname]
    dv = DataValidation(type="list",formula1=f"={src}",allow_blank=True,showDropDown=False)
    dv.error = "Pick a value from the dropdown"
    dv.errorTitle = "Invalid value"
    dv.prompt = f"Choose: {label}"
    dv.add(f"{col}2:{col}{sheet_last[sname]}")
    target.add_data_validation(dv)

wb._sheets = [wb[n] for n in ["README","Page Inventory","Keyword Targets","Technical Health","Internal Linking","Performance","Backlinks","Roadmap","Lookups"]]
wb.save(OUT)
print(f"✓ Saved: {OUT}")
